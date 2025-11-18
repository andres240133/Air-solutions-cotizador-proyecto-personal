"""
excel_manager.py - Gestor de Exportación/Importación de Excel

Permite exportar proyectos a Excel en formato editable e importar cambios
"""

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelManager:
    """Gestor de Excel para proyectos"""

    def __init__(self, db):
        """
        Inicializa el gestor

        Args:
            db: Instancia de DatabaseManager
        """
        self.db = db

    def exportar_proyecto(self, id_proyecto):
        """
        Exporta un proyecto a Excel

        Args:
            id_proyecto: ID del proyecto a exportar

        Returns:
            str: Ruta del archivo creado
        """
        # Obtener datos del proyecto
        cursor = self.db.ejecutar_query('''
            SELECT
                p.numero_proyecto, p.nombre_proyecto,
                c.nombre_empresa, p.ubicacion, p.responsable,
                p.fecha_inicio, p.descripcion
            FROM proyectos p
            LEFT JOIN clientes c ON p.id_cliente = c.id_cliente
            WHERE p.id_proyecto = ?
        ''', (id_proyecto,))

        proyecto = cursor.fetchone()

        if not proyecto:
            raise Exception("Proyecto no encontrado")

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Proyecto"

        # Estilos
        header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        subheader_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
        subheader_font = Font(bold=True, color="FFFFFF", size=11)
        nivel_fill = PatternFill(start_color="e5e7eb", end_color="e5e7eb", fill_type="solid")
        nivel_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
        total_font = Font(bold=True, color="FFFFFF", size=11)

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # === ENCABEZADO DEL PROYECTO ===
        row = 1

        # Título principal
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = f"{proyecto[0]} - {proyecto[1]}"
        cell.font = Font(bold=True, size=16, color="1e293b")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

        # Información del proyecto
        ws.merge_cells(f'A{row}:I{row}')
        info_text = f"Cliente: {proyecto[2] or 'N/A'}"
        if proyecto[3]:
            info_text += f"  |  Ubicación: {proyecto[3]}"
        if proyecto[4]:
            info_text += f"  |  Responsable: {proyecto[4]}"
        ws[f'A{row}'].value = info_text
        ws[f'A{row}'].font = Font(size=10, italic=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1

        if proyecto[6]:
            ws.merge_cells(f'A{row}:I{row}')
            ws[f'A{row}'].value = f"Descripción: {proyecto[6]}"
            ws[f'A{row}'].font = Font(size=9, italic=True)
            row += 1

        row += 1  # Espacio

        # === ENCABEZADOS DE COLUMNAS ===
        headers = [
            'Nivel', 'Especificación', 'Descripción', 'Cantidad', 'Unidad',
            'Costo Equipo', 'Costo Materiales', 'Costo Mano Obra', 'Total'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        row += 1

        # === DATOS POR NIVEL ===
        cursor = self.db.ejecutar_query('''
            SELECT id_nivel, codigo_nivel, nombre_nivel, total_nivel
            FROM proyecto_niveles
            WHERE id_proyecto = ?
            ORDER BY orden, codigo_nivel
        ''', (id_proyecto,))

        niveles = cursor.fetchall()

        total_equipos_proyecto = 0
        total_materiales_proyecto = 0
        total_mano_obra_proyecto = 0
        total_proyecto = 0

        for nivel in niveles:
            id_nivel, codigo_nivel, nombre_nivel, total_nivel = nivel

            # Fila del nivel
            ws.merge_cells(f'A{row}:C{row}')
            cell = ws[f'A{row}']
            cell.value = f"{codigo_nivel} - {nombre_nivel}"
            cell.fill = nivel_fill
            cell.font = nivel_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border

            # Total del nivel
            ws.merge_cells(f'D{row}:H{row}')
            ws[f'D{row}'].border = border

            cell = ws[f'I{row}']
            cell.value = total_nivel or 0
            cell.number_format = '"$"#,##0.00'
            cell.fill = nivel_fill
            cell.font = nivel_font
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = border

            row += 1

            # Items del nivel
            cursor_items = self.db.ejecutar_query('''
                SELECT
                    especificacion, descripcion, cantidad, unidad,
                    costo_equipo, costo_materiales, costo_mano_obra,
                    total_item
                FROM proyecto_items
                WHERE id_nivel = ?
                ORDER BY orden, especificacion
            ''', (id_nivel,))

            items = cursor_items.fetchall()

            for item in items:
                (espec, desc, cant, unidad, c_equipo, c_materiales,
                 c_mano_obra, total_item) = item

                # Código nivel (vacío para items)
                ws[f'A{row}'].border = border

                # Especificación
                cell = ws[f'B{row}']
                cell.value = espec
                cell.border = border

                # Descripción
                cell = ws[f'C{row}']
                cell.value = desc or ''
                cell.border = border

                # Cantidad
                cell = ws[f'D{row}']
                cell.value = cant
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='right')
                cell.border = border

                # Unidad
                cell = ws[f'E{row}']
                cell.value = unidad
                cell.alignment = Alignment(horizontal='center')
                cell.border = border

                # Costo Equipo
                cell = ws[f'F{row}']
                cell.value = c_equipo
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal='right')
                cell.border = border

                # Costo Materiales
                cell = ws[f'G{row}']
                cell.value = c_materiales
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal='right')
                cell.border = border

                # Costo Mano de Obra
                cell = ws[f'H{row}']
                cell.value = c_mano_obra
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal='right')
                cell.border = border

                # Total
                cell = ws[f'I{row}']
                cell.value = total_item
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal='right')
                cell.border = border

                # Acumular totales
                total_equipos_proyecto += c_equipo * cant
                total_materiales_proyecto += c_materiales * cant
                total_mano_obra_proyecto += c_mano_obra * cant

                row += 1

            total_proyecto += total_nivel or 0
            row += 1  # Espacio entre niveles

        # === TOTALES GENERALES ===
        row += 1

        # Total Equipos
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = "TOTAL EQUIPOS"
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border

        ws.merge_cells(f'F{row}:I{row}')
        cell = ws[f'F{row}']
        cell.value = total_equipos_proyecto
        cell.number_format = '"$"#,##0.00'
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border

        row += 1

        # Total Materiales
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = "TOTAL MATERIALES"
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border

        ws.merge_cells(f'F{row}:I{row}')
        cell = ws[f'F{row}']
        cell.value = total_materiales_proyecto
        cell.number_format = '"$"#,##0.00'
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border

        row += 1

        # Total Mano de Obra
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = "TOTAL MANO DE OBRA"
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border

        ws.merge_cells(f'F{row}:I{row}')
        cell = ws[f'F{row}']
        cell.value = total_mano_obra_proyecto
        cell.number_format = '"$"#,##0.00'
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border

        row += 1

        # Total General
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = "TOTAL GENERAL DEL PROYECTO"
        cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border

        ws.merge_cells(f'F{row}:I{row}')
        cell = ws[f'F{row}']
        cell.value = total_proyecto
        cell.number_format = '"$"#,##0.00'
        cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 15

        # Guardar archivo
        carpeta_exports = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            'exports'
        )
        os.makedirs(carpeta_exports, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"{proyecto[0]}__{timestamp}.xlsx"
        ruta_archivo = os.path.join(carpeta_exports, nombre_archivo)

        wb.save(ruta_archivo)

        return ruta_archivo

    def importar_proyecto(self, id_proyecto, ruta_excel):
        """
        Importa cambios desde Excel

        Args:
            id_proyecto: ID del proyecto
            ruta_excel: Ruta del archivo Excel

        Returns:
            tuple: (exito, mensaje)
        """
        try:
            wb = load_workbook(ruta_excel)
            ws = wb.active

            # Verificar formato
            if ws['B5'].value != 'Especificación':
                return (False, "El archivo no tiene el formato esperado")

            # Obtener niveles del proyecto
            cursor = self.db.ejecutar_query('''
                SELECT id_nivel, codigo_nivel
                FROM proyecto_niveles
                WHERE id_proyecto = ?
                ORDER BY orden, codigo_nivel
            ''', (id_proyecto,))

            niveles_dict = {codigo: id_nivel for id_nivel, codigo in cursor.fetchall()}

            # Leer datos del Excel
            row = 6  # Comienza después del header
            nivel_actual = None
            items_actualizados = 0
            items_nuevos = 0

            while row <= ws.max_row:
                # Verificar si es una fila de nivel
                nivel_codigo = ws[f'A{row}'].value

                if nivel_codigo and '-' in str(nivel_codigo):
                    # Es un nivel
                    codigo_nivel = str(nivel_codigo).split('-')[0].strip()
                    if codigo_nivel in niveles_dict:
                        nivel_actual = niveles_dict[codigo_nivel]
                    row += 1
                    continue

                # Es un item
                if nivel_actual and ws[f'B{row}'].value:
                    especificacion = ws[f'B{row}'].value
                    descripcion = ws[f'C{row}'].value
                    cantidad = ws[f'D{row}'].value or 1
                    unidad = ws[f'E{row}'].value or 'unidad'
                    costo_equipo = ws[f'F{row}'].value or 0
                    costo_materiales = ws[f'G{row}'].value or 0
                    costo_mano_obra = ws[f'H{row}'].value or 0

                    # Verificar si el item existe
                    cursor_check = self.db.ejecutar_query('''
                        SELECT id_item
                        FROM proyecto_items
                        WHERE id_nivel = ? AND especificacion = ?
                    ''', (nivel_actual, especificacion))

                    item_existe = cursor_check.fetchone()

                    if item_existe:
                        # Actualizar item existente
                        total_item = (costo_equipo + costo_materiales + costo_mano_obra) * cantidad

                        self.db.ejecutar_query('''
                            UPDATE proyecto_items
                            SET descripcion = ?, cantidad = ?, unidad = ?,
                                costo_equipo = ?, costo_materiales = ?, costo_mano_obra = ?,
                                total_item = ?
                            WHERE id_item = ?
                        ''', (
                            descripcion, cantidad, unidad,
                            costo_equipo, costo_materiales, costo_mano_obra,
                            total_item, item_existe[0]
                        ))
                        items_actualizados += 1
                    else:
                        # Crear nuevo item
                        total_item = (costo_equipo + costo_materiales + costo_mano_obra) * cantidad

                        self.db.ejecutar_query('''
                            INSERT INTO proyecto_items (
                                id_nivel, especificacion, descripcion,
                                cantidad, unidad,
                                costo_equipo, costo_materiales, costo_mano_obra,
                                total_item, orden
                            )
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 999)
                        ''', (
                            nivel_actual, especificacion, descripcion,
                            cantidad, unidad,
                            costo_equipo, costo_materiales, costo_mano_obra,
                            total_item
                        ))
                        items_nuevos += 1

                row += 1

            mensaje = f"Importación exitosa:\n"
            mensaje += f"- {items_actualizados} items actualizados\n"
            mensaje += f"- {items_nuevos} items nuevos agregados"

            return (True, mensaje)

        except Exception as e:
            return (False, f"Error al importar Excel:\n{str(e)}")
